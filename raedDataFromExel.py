import numpy as np
from openpyxl import load_workbook
from sklearn.model_selection import train_test_split
myFileName= r'C:/Users/shayanr/Desktop/tehran/lam/data/FV.xlsx'
#load the workbook, and put the sheet into a variable
wb = load_workbook(filename=myFileName)
ws = wb['1']
maxRow = ws.max_row +1
X=[]
Y=[]
for i in range(maxRow-2):
    x=[]
    y=[]
    if(ws.cell(i + 2,1).value!=2):
     y.append(1)
    else:
        y.append(ws.cell(i + 2,1).value)
    for j in range(128):
     x.append(ws.cell(i+2,j+2).value)
    X.append(x)
    Y.append(y)
X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.10, random_state=42)
from sklearn.svm import SVC # "Support vector classifier"
model = SVC(kernel='linear', C=1E10)
model.fit(X_train, y_train)
