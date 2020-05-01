from openpyxl import load_workbook
import cv2
import numpy as np
from closestPoint import closestPoint

myFileName=r'C:/Users/shayanr/Desktop/tehran/lam/data/label.xlsx'
#load the workbook, and put the sheet into a variable
wb = load_workbook(filename=myFileName)
ws = wb['1']

#max_row is a sheet function that gets the last row in a sheet.
maxRow = ws.max_row +1
pic=""
keypoints=[]
descriptors=[]
for i in range(maxRow):
    if ws.cell(i+1,4).value!=pic :
        if descriptors!=[]:
            idx = np.random.randint(0, descriptors.shape[0],20)
            descriptorsX = descriptors[idx]

            for m in range(len(descriptorsX)):
                myFileName2 = r'C:/Users/shayanr/Desktop/tehran/lam/data/FV.xlsx'
                # load the workbook, and put the sheet into a variable
                wb2 = load_workbook(filename=myFileName2)
                ws2 = wb2['1']
                newRowLocation = ws2.max_row + 1
                ws2.cell(column=1, row=newRowLocation, value=6)
                for j in range(len(descriptors[m])):
                    ws2.cell(column=j + 2, row=newRowLocation, value=descriptorsX[m][j])
                wb2.save(filename=myFileName2)
                wb2.close()
            cv2.imwrite('C:/Users/shayanr/Desktop/tehran/lam/data/labeld/' +pic, image)
        pic=ws.cell(i+1,4).value
        print(pic)
        image = cv2.imread('C:/Users/shayanr/Desktop/tehran/lam/data/allData/'+str(pic))
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        # DOG_thresh=0.001,corn_thresh=0.000001
        harris = cv2.xfeatures2d.HarrisLaplaceFeatureDetector_create(DOG_thresh=0.001, corn_thresh=0.00001)
        # Detect key points
        keypoints = harris.detect(gray, None)
        #freakExtractor = cv2.xfeatures2d.FREAK_create()
        sift = cv2.xfeatures2d.SIFT_create(edgeThreshold=500, contrastThreshold=0.01)
        keypoints, descriptors = sift.compute(gray, keypoints)

    for x in range(1):
        (keypoint, descriptor) = closestPoint(keypoints, descriptors, ws.cell(i+1,2).value, ws.cell(i+1,3).value)
        image = cv2.circle(image, (int(keypoint.pt[0]), int(keypoint.pt[1])), 10, (0, 0, 255), -1)
        myFileName2 =  r'C:/Users/shayanr/Desktop/tehran/lam/data/FV.xlsx'
        # load the workbook, and put the sheet into a variable
        wb2 = load_workbook(filename=myFileName2)
        ws2 = wb2['1']
        newRowLocation = ws2.max_row + 1
        ws2.cell(column=1, row=newRowLocation, value=ws.cell(i+1,1).value)
        for j in range(len(descriptor)):
            ws2.cell(column=j+2, row=newRowLocation, value=descriptor[j])
        index=keypoints.index(keypoint)
        keypoints.remove(keypoint)
        descriptors = np.delete(descriptors, index, axis=0)
        wb2.save(filename=myFileName2)
        wb2.close()


